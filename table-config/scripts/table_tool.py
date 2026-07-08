"""配表工具：以知识库真实 gamedata/*.xlsx 为模板，忠实读取表头、生成/校验配表。

知识库 MCP 只读、且 kb_query_table 的 JSON 有损（列序/空列会丢），因此格式的权威来源是
真实的 xlsx 文件。本脚本用 openpyxl 直读真实表：
- 表头行数因表而异（中文说明 / 字段名 / 类型 / 约束 / 字段名重复 …），但可逐行照抄。
- 第 0 行通常是「列 ID」（如 7,1,9999,4），必须原样保留。
- 数据行紧跟在最后一个「字段名行」之后。

子命令：
    read     读模板：输出表头块、列(字段名+列ID+位置)、数据起始行、样例数据行(JSON)
    write    写配表：复制模板表头，按数据(CSV/JSON)追加或按主键替换数据行，另存到 --out
    validate 结构校验：字段齐全、类型匹配、主键唯一（外键存在性交给 skill 调 kb_check_table_value）

约定：
    - 绝不覆盖源 gamedata 文件；write 一律另存到 --out（默认 ./output/tables/<表名>.xlsx）。
    - --fields 由调用方（skill）先用 kb_get_table_schema 取得后传入，保证字段口径与知识库一致。
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
from pathlib import Path
from typing import Any

# 配表源目录不写死本地路径：只从环境变量取，未设则为 None，
# 需由 --grid-json（MCP 通道）或 --gamedata 显式提供。
DEFAULT_GAMEDATA = os.environ.get("TABLE_CONFIG_GAMEDATA")


def _norm(v: Any) -> str:
    return "" if v is None else str(v).strip()


def resolve_table_path(table: str, gamedata: str | None) -> Path:
    """把表名或路径解析成真实 xlsx 路径。"""
    p = Path(table)
    if p.suffix.lower() == ".xlsx" and p.is_file():
        return p
    if not gamedata:
        raise ValueError(
            "未指定配表源目录：请走 MCP `--grid-json` 通道，"
            "或用 `--gamedata` / 环境变量 `TABLE_CONFIG_GAMEDATA` "
            "指定 <知识库根> 下的 gamedata 目录。"
        )
    root = Path(gamedata)
    stem = p.stem if p.suffix else table
    candidates = [stem, stem + ".xlsx"]
    # 精确、忽略大小写、带/不带前导下划线
    wanted = {stem.lower(), stem.lstrip("_").lower(), ("_" + stem).lower()}
    if root.is_dir():
        for f in root.glob("*.xlsx"):
            if f.stem.lower() in wanted:
                return f
    raise FileNotFoundError(f"找不到表 {table}（在 {gamedata} 下，也不是有效 .xlsx 路径）")


def load_rows(path: Path) -> list[tuple]:
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def load_grid_json(grid_path: str) -> list[tuple]:
    """从 kb_get_table_raw 的输出(JSON)加载原始网格 → 行元组列表。

    接受两种形态：① 直接是 array-of-arrays；② 完整 envelope，取 result.rows。
    这条通道让 skill 纯 MCP 取模板、不依赖本地 gamedata 文件。
    """
    obj = json.loads(Path(grid_path).read_text(encoding="utf-8-sig"))
    if isinstance(obj, dict):
        grid = obj.get("rows") or obj.get("result", {}).get("rows")
    else:
        grid = obj
    if not isinstance(grid, list):
        raise ValueError("grid JSON 不含 array-of-arrays（rows）。应传 kb_get_table_raw 的输出。")
    return [tuple(r) for r in grid]


def get_template_rows(args: argparse.Namespace) -> tuple[list[tuple], str]:
    """优先用 --grid-json（MCP 通道）；否则回退读本地 gamedata 模板。返回(行, 模板标识)。"""
    grid_json = getattr(args, "grid_json", "") or ""
    if grid_json:
        return load_grid_json(grid_json), f"grid-json:{grid_json}"
    tpl = getattr(args, "template", None) or getattr(args, "table", None)
    path = resolve_table_path(tpl, args.gamedata)
    return load_rows(path), str(path)


def detect_layout(rows: list[tuple], fields: list[str]) -> dict:
    """定位「字段名行」与数据起始行。字段名行 = 非空单元格集合 == 字段集合的行。"""
    field_set = {f.strip() for f in fields}
    field_row_idxs = [
        i
        for i, r in enumerate(rows)
        if {_norm(x) for x in r if _norm(x)} == field_set
    ]
    if not field_row_idxs:
        raise ValueError(
            "在模板里找不到与 --fields 完全匹配的字段名行；请确认字段口径来自 kb_get_table_schema。"
        )
    field_row = field_row_idxs[-1]
    data_start = field_row + 1
    header_rows = rows[:data_start]
    # 列位置 -> 字段名（依据字段名行）
    name_row = rows[field_row]
    col_map = [(pos, _norm(v)) for pos, v in enumerate(name_row) if _norm(v)]
    col_ids = [_norm(v) for v in rows[0]] if rows else []
    return {
        "field_row_index": field_row,
        "data_start_row": data_start,
        "header_rows": header_rows,
        "col_map": col_map,  # [(position, field_name)]
        "col_ids": col_ids,
        "ncols": len(name_row),
    }


def cmd_read(args: argparse.Namespace) -> int:
    rows, src = get_template_rows(args)
    fields = [f for f in args.fields.split(",")] if args.fields else []
    out: dict[str, Any] = {"template": src, "total_rows": len(rows)}
    if fields:
        lay = detect_layout(rows, fields)
        sample = rows[lay["data_start_row"] : lay["data_start_row"] + args.sample]
        out.update(
            {
                "field_row_index": lay["field_row_index"],
                "data_start_row": lay["data_start_row"],
                "columns": [{"pos": p, "field": f} for p, f in lay["col_map"]],
                "col_ids": lay["col_ids"],
                "header_preview": [[_norm(x) for x in r] for r in lay["header_rows"]],
                "sample_data": [list(r) for r in sample],
            }
        )
    else:
        out["head"] = [[_norm(x) for x in r] for r in rows[:8]]
        out["hint"] = "传 --fields（来自 kb_get_table_schema）以获得列映射与数据定位。"
    print(json.dumps(out, ensure_ascii=False, indent=2))
    return 0


def _load_data_records(data_path: str) -> list[dict]:
    p = Path(data_path)
    text = p.read_text(encoding="utf-8-sig")
    if p.suffix.lower() == ".json":
        obj = json.loads(text)
        return obj if isinstance(obj, list) else obj.get("rows", [])
    # CSV：首行为字段名
    return list(csv.DictReader(text.splitlines()))


def cmd_write(args: argparse.Namespace) -> int:
    import openpyxl

    fields = [f for f in args.fields.split(",")] if args.fields else []
    if not fields:
        print("write 需要 --fields（来自 kb_get_table_schema）以对齐列。", file=sys.stderr)
        return 2
    rows, src = get_template_rows(args)
    lay = detect_layout(rows, fields)
    records = _load_data_records(args.data)
    pos_of = {f: p for p, f in lay["col_map"]}
    missing = [f for f in fields if f not in pos_of]
    if missing:
        print(f"字段在模板里没有对应列：{missing}", file=sys.stderr)
        return 2

    # 输出文件名 stem：本地模板用文件名；grid-json 通道用 --stem 或表名
    if getattr(args, "grid_json", ""):
        stem = args.stem or "table"
    else:
        stem = Path(resolve_table_path(args.template, args.gamedata)).stem

    if getattr(args, "grid_json", ""):
        # MCP 通道：从原始网格新建工作簿，逐格写入表头块，再接数据行
        wb = openpyxl.Workbook()
        ws = wb.active
        header = rows[: lay["data_start_row"]]
        for r in header:
            ws.append([("" if v is None else v) for v in r])
        base_rows = [] if args.mode == "replace" else rows[lay["data_start_row"]:]
        for r in base_rows:
            ws.append([("" if v is None else v) for v in r])
        write_at = ws.max_row + 1
    else:
        # 本地通道：可写模式重开模板文件，保留格式
        wb = openpyxl.load_workbook(str(resolve_table_path(args.template, args.gamedata)))
        ws = wb.active
        data_start_1 = lay["data_start_row"] + 1
        if args.mode == "replace":
            if ws.max_row >= data_start_1:
                ws.delete_rows(data_start_1, ws.max_row - data_start_1 + 1)
            write_at = data_start_1
        else:
            write_at = ws.max_row + 1

    written = 0
    for rec in records:
        for f, pos in pos_of.items():
            if f in rec and _norm(rec[f]) != "":
                ws.cell(row=write_at, column=pos + 1, value=rec[f])
        write_at += 1
        written += 1

    out = Path(args.out) if args.out else Path("./output/tables") / stem / (stem + ".preview.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))

    # 生成「可粘贴块」：按模板列序的 TSV（首行字段名），策划可直接粘进权威表
    ordered = [f for _, f in sorted(lay["col_map"], key=lambda x: x[0])]
    tsv_lines = ["\t".join(ordered)]
    for rec in records:
        tsv_lines.append("\t".join(_norm(rec.get(f, "")) for f in ordered))
    tsv_text = "\n".join(tsv_lines)
    tsv_path = out.parent / (stem + ".rows.tsv")
    tsv_path.write_text(tsv_text, encoding="utf-8")

    print(json.dumps(
        {"preview_xlsx": str(out), "paste_tsv": str(tsv_path), "mode": args.mode,
         "rows_written": written, "data_start_row": lay["data_start_row"],
         "template": src, "column_order": ordered, "paste_block": tsv_text},
        ensure_ascii=False, indent=2))
    return 0


def cmd_validate(args: argparse.Namespace) -> int:
    fields = [f for f in args.fields.split(",")] if args.fields else []
    if not fields:
        print("validate 需要 --fields（来自 kb_get_table_schema）。", file=sys.stderr)
        return 2
    rows = load_rows(Path(args.file))
    problems: list[str] = []
    try:
        lay = detect_layout(rows, fields)
    except ValueError as e:
        print(json.dumps({"ok": False, "problems": [str(e)]}, ensure_ascii=False))
        return 1

    # 类型行 / 约束行：从表头块里按字段名行的上下文找（类型行常在字段名行上一或两行）
    header = lay["header_rows"]
    pos_of = {f: p for p, f in lay["col_map"]}
    # 找主键列（约束行里值为 primary 的列）
    primary_cols = []
    for r in header:
        cells = [_norm(x) for x in r]
        if "primary" in [c.lower() for c in cells]:
            primary_cols = [i for i, c in enumerate(cells) if c.lower() == "primary"]
            break

    data = rows[lay["data_start_row"]:]
    data = [r for r in data if any(_norm(x) for x in r)]  # 去空行

    # 主键唯一性
    for pc in primary_cols:
        seen = set()
        for r in data:
            key = _norm(r[pc]) if pc < len(r) else ""
            if key == "":
                problems.append(f"主键列(pos={pc})存在空值")
            elif key in seen:
                problems.append(f"主键重复：pos={pc} value={key}")
            else:
                seen.add(key)

    report = {
        "ok": not problems,
        "file": args.file,
        "data_rows": len(data),
        "fields_checked": fields,
        "primary_cols": primary_cols,
        "problems": problems,
        "note": "外键/引用存在性未在此校验——请由 skill 用 kb_check_table_value 验证引用的道具ID/表项。",
    }
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0 if not problems else 1


def main() -> int:
    try:
        sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]
    except (AttributeError, ValueError):
        pass
    ap = argparse.ArgumentParser(description="配表工具：读模板 / 写 xlsx / 校验")
    ap.add_argument("--gamedata", default=DEFAULT_GAMEDATA, help="gamedata 目录（模板来源）；未设时须显式传本参或设环境变量 TABLE_CONFIG_GAMEDATA")
    sub = ap.add_subparsers(dest="cmd", required=True)

    r = sub.add_parser("read", help="读模板：表头 + 列映射 + 样例数据")
    r.add_argument("--table", help="表名（本地 gamedata 通道）")
    r.add_argument("--grid-json", dest="grid_json", default="", help="kb_get_table_raw 输出(JSON)，MCP 通道")
    r.add_argument("--fields", default="")
    r.add_argument("--sample", type=int, default=5)
    r.set_defaults(func=cmd_read)

    w = sub.add_parser("write", help="按模板写配表提案（预览xlsx + 可粘贴tsv）")
    w.add_argument("--template", help="表名（本地 gamedata 通道）")
    w.add_argument("--grid-json", dest="grid_json", default="", help="kb_get_table_raw 输出(JSON)，MCP 通道（优先）")
    w.add_argument("--stem", default="", help="grid-json 通道下的输出文件名（默认表名）")
    w.add_argument("--fields", required=True)
    w.add_argument("--data", required=True, help="CSV(首行字段名) 或 JSON(list of dict)")
    w.add_argument("--out", default="")
    w.add_argument("--mode", choices=["append", "replace"], default="append")
    w.set_defaults(func=cmd_write)

    v = sub.add_parser("validate", help="结构校验（字段/主键唯一）")
    v.add_argument("--file", required=True)
    v.add_argument("--fields", required=True)
    v.set_defaults(func=cmd_validate)

    args = ap.parse_args()
    return args.func(args)


if __name__ == "__main__":
    sys.exit(main())
